# Create your views here.
"""
Vistas para el módulo de reportes NEXO
"""
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
import logging
import time
from datetime import datetime

# Importar servicios y formularios
from .services import ReportService
from .forms import (
    InventarioGeneralForm, ProduccionForm, VentasForm, DevolucionesForm,
    ClientesForm, UsuariosEmpleadosForm, ProductosCategoriaForm, AuditoriaForm
)

# Importar decoradores de autenticación
try:
    from AuthLogin.decorators import nexo_login_required, nexo_role_required
except ImportError:
    def nexo_login_required(view_func):
        return view_func
    
    def nexo_role_required(roles):
        def decorator(view_func):
            return view_func
        return decorator

logger = logging.getLogger('nexo.Informes')

# Configuración de colores NEXO
COLORES_NEXO = {
    'primary': '#4ECDC4',
    'secondary_yellow': '#F2CE16', 
    'accent_orange': '#F29D35',
    'accent_dark_orange': '#F28627',
    'bg_light': '#eaeef3',
    'bg_very_light': '#F2F2F2',
    'white': '#ffffff',
    'text_dark': '#374151',
    'text_medium': '#6b7280',
    'text_light': '#9ca3af',
    'success': '#4ECDC4',
    'warning': '#F2CE16',
    'danger': '#F28627',
    'info': '#2752F2'
}

def get_client_ip(request):
    """Obtiene la IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@nexo_login_required
def reports_dashboard(request):
    """
    Dashboard principal de reportes
    """
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    # Definir reportes disponibles según el rol
    reportes_disponibles = []
    
    if nexo_user_role == 'admin':
        reportes_disponibles = [
            {
                'nombre': 'Inventario General',
                'descripcion': 'Existencias actuales por producto y ubicación',
                'url': 'Informes:inventario_general',
                'icon': 'fas fa-boxes',
                'color': 'bg-blue-500'
            },
            {
                'nombre': 'Producción',
                'descripcion': 'Productos elaborados en taller, fechas y responsables',
                'url': 'Informes:produccion',
                'icon': 'fas fa-industry',
                'color': 'bg-green-500'
            },
            {
                'nombre': 'Reabastecimientos',
                'descripcion': 'Productos transferidos desde taller a sucursal',
                'url': 'Informes:reabastecimientos',
                'icon': 'fas fa-truck',
                'color': 'bg-yellow-500'
            },
            {
                'nombre': 'Ventas',
                'descripcion': 'Registros de ventas por cliente, fecha y vendedor',
                'url': 'Informes:ventas',
                'icon': 'fas fa-shopping-cart',
                'color': 'bg-purple-500'
            },
            {
                'nombre': 'Devoluciones',
                'descripcion': 'Devoluciones procesadas, causas y productos',
                'url': 'Informes:devoluciones',
                'icon': 'fas fa-undo',
                'color': 'bg-red-500'
            },
            {
                'nombre': 'Clientes',
                'descripcion': 'Estado actual de los clientes del sistema',
                'url': 'Informes:clientes',
                'icon': 'fas fa-users',
                'color': 'bg-indigo-500'
            },
            {
                'nombre': 'Usuarios y Empleados',
                'descripcion': 'Roles, accesos y personal activo',
                'url': 'Informes:usuarios_empleados',
                'icon': 'fas fa-user-tie',
                'color': 'bg-gray-500'
            },
            {
                'nombre': 'Productos por Categoría',
                'descripcion': 'Distribución y rotación por categoría',
                'url': 'Informes:productos_categoria',
                'icon': 'fas fa-tags',
                'color': 'bg-pink-500'
            },
            {
                'nombre': 'Auditoría del Sistema',
                'descripcion': 'Eventos clave: registros, ediciones, eliminaciones',
                'url': 'Informes:auditoria',
                'icon': 'fas fa-shield-alt',
                'color': 'bg-orange-500'
            }
        ]
    elif nexo_user_role == 'encargado_sucursal':
        reportes_disponibles = [
            {
                'nombre': 'Inventario General',
                'descripcion': 'Existencias de sucursal únicamente',
                'url': 'Informes:inventario_general',
                'icon': 'fas fa-boxes',
                'color': 'bg-blue-500'
            },
            {
                'nombre': 'Ventas',
                'descripcion': 'Registros de ventas de la sucursal',
                'url': 'Informes:ventas',
                'icon': 'fas fa-shopping-cart',
                'color': 'bg-purple-500'
            },
            {
                'nombre': 'Devoluciones',
                'descripcion': 'Devoluciones de la sucursal',
                'url': 'Informes:devoluciones',
                'icon': 'fas fa-undo',
                'color': 'bg-red-500'
            },
            {
                'nombre': 'Clientes',
                'descripcion': 'Clientes de la sucursal',
                'url': 'Informes:clientes',
                'icon': 'fas fa-users',
                'color': 'bg-indigo-500'
            }
        ]
    
    context = {
        'reportes_disponibles': reportes_disponibles,
        'page_title': 'Centro de Reportes',
        'page_subtitle': 'Genera reportes detallados del sistema NEXO',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/dashboard.html', context)

@nexo_login_required
@nexo_role_required(['admin', 'encargado_sucursal'])
def inventario_general(request):
    """
    Reporte de inventario general
    """
    start_time = time.time()
    form = InventarioGeneralForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        # Aplicar filtros según el rol
        usuario_actual = getattr(request, 'nexo_user', None)
        if usuario_actual and usuario_actual.rol == 'encargado_sucursal':
            # Solo inventario de sucursal (ubicación 2)
            filtros['ubicacion'] = 2
        else:
            # Admin puede ver todas las ubicaciones
            if form.cleaned_data.get('ubicacion'):
                filtros['ubicacion'] = form.cleaned_data['ubicacion'].id_ubicacion
        
        if form.cleaned_data.get('categoria'):
            filtros['categoria'] = form.cleaned_data['categoria'].idcategoria
        
        if form.cleaned_data.get('stock_bajo'):
            filtros['stock_bajo'] = True
        
        # Obtener datos
        resultado = ReportService.get_inventario_general(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='inventario_general',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Inventario General', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Inventario General', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Inventario General',
        'page_subtitle': 'Existencias actuales por producto y ubicación',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/inventario_general.html', context)

@nexo_login_required
@nexo_role_required(['admin'])
def produccion(request):
    """
    Reporte de producción
    """
    start_time = time.time()
    form = ProduccionForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('fecha_desde'):
            filtros['fecha_desde'] = form.cleaned_data['fecha_desde']
        
        if form.cleaned_data.get('fecha_hasta'):
            filtros['fecha_hasta'] = form.cleaned_data['fecha_hasta']
        
        if form.cleaned_data.get('usuario'):
            filtros['usuario'] = form.cleaned_data['usuario'].idusuario
        
        # Obtener datos
        resultado = ReportService.get_produccion(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='produccion',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Producción', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Producción', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Producción',
        'page_subtitle': 'Productos elaborados en taller, fechas y responsables',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/produccion.html', context)

@nexo_login_required
@nexo_role_required(['admin'])
def reabastecimientos(request):
    """
    Reporte de reabastecimientos a sucursal
    """
    # Por implementar - requiere lógica de transferencias
    messages.info(request, 'Reporte de reabastecimientos en desarrollo.')
    return redirect('Informes:Informes_dashboard')

@nexo_login_required
@nexo_role_required(['admin', 'encargado_sucursal'])
def ventas(request):
    """
    Reporte de ventas
    """
    start_time = time.time()
    form = VentasForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('fecha_desde'):
            filtros['fecha_desde'] = form.cleaned_data['fecha_desde']
        
        if form.cleaned_data.get('fecha_hasta'):
            filtros['fecha_hasta'] = form.cleaned_data['fecha_hasta']
        
        if form.cleaned_data.get('cliente'):
            filtros['cliente'] = form.cleaned_data['cliente'].idcliente
        
        if form.cleaned_data.get('vendedor'):
            filtros['vendedor'] = form.cleaned_data['vendedor'].idusuario
        
        if form.cleaned_data.get('estado'):
            filtros['estado'] = form.cleaned_data['estado']
        
        # Obtener datos
        resultado = ReportService.get_ventas(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='ventas',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Ventas', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Ventas', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Ventas',
        'page_subtitle': 'Registros de ventas por cliente, fecha y vendedor',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/ventas.html', context)

@nexo_login_required
@nexo_role_required(['admin', 'encargado_sucursal'])
def devoluciones(request):
    """
    Reporte de devoluciones
    """
    start_time = time.time()
    form = DevolucionesForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('fecha_desde'):
            filtros['fecha_desde'] = form.cleaned_data['fecha_desde']
        
        if form.cleaned_data.get('fecha_hasta'):
            filtros['fecha_hasta'] = form.cleaned_data['fecha_hasta']
        
        if form.cleaned_data.get('motivo'):
            filtros['motivo'] = form.cleaned_data['motivo']
        
        # Obtener datos
        resultado = ReportService.get_devoluciones(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='devoluciones',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Devoluciones', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Devoluciones', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Devoluciones',
        'page_subtitle': 'Devoluciones procesadas, causas y productos implicados',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/devoluciones.html', context)

@nexo_login_required
@nexo_role_required(['admin', 'encargado_sucursal'])
def clientes(request):
    """
    Reporte de clientes
    """
    start_time = time.time()
    form = ClientesForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('estado') != '':
            filtros['estado'] = form.cleaned_data['estado']
        
        # Obtener datos
        resultado = ReportService.get_clientes(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='clientes',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Clientes', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Clientes', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Clientes',
        'page_subtitle': 'Estado actual de los clientes del sistema',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/clientes.html', context)

@nexo_login_required
@nexo_role_required(['admin'])
def usuarios_empleados(request):
    """
    Reporte de usuarios y empleados
    """
    start_time = time.time()
    form = UsuariosEmpleadosForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('rol'):
            filtros['rol'] = form.cleaned_data['rol']
        
        if form.cleaned_data.get('estado') != '':
            filtros['estado'] = form.cleaned_data['estado']
        
        # Obtener datos
        resultado = ReportService.get_usuarios_empleados(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='usuarios_empleados',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Usuarios y Empleados', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Usuarios y Empleados', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Usuarios y Empleados',
        'page_subtitle': 'Roles, accesos y relación con personal activo',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/usuarios_empleados.html', context)

@nexo_login_required
@nexo_role_required(['admin'])
def productos_categoria(request):
    """
    Reporte de productos por categoría
    """
    start_time = time.time()
    form = ProductosCategoriaForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('fecha_desde'):
            filtros['fecha_desde'] = form.cleaned_data['fecha_desde']
        
        if form.cleaned_data.get('fecha_hasta'):
            filtros['fecha_hasta'] = form.cleaned_data['fecha_hasta']
        
        if form.cleaned_data.get('categoria'):
            filtros['categoria'] = form.cleaned_data['categoria'].idcategoria
        
        # Obtener datos
        resultado = ReportService.get_productos_categoria(filtros)
        data = resultado.get('data', [])
        stats = resultado.get('stats', {})
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='productos_categoria',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Productos por Categoría', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Productos por Categoría', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Productos por Categoría',
        'page_subtitle': 'Distribución y rotación de productos según categoría',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/productos_categoria.html', context)

@nexo_login_required
@nexo_role_required(['admin'])
def auditoria(request):
    """
    Reporte de auditoría del sistema
    """
    start_time = time.time()
    form = AuditoriaForm(request.GET or None)
    data = []
    stats = {}
    
    if form.is_valid():
        filtros = {}
        
        if form.cleaned_data.get('fecha_desde'):
            filtros['fecha_desde'] = form.cleaned_data['fecha_desde']
        
        if form.cleaned_data.get('fecha_hasta'):
            filtros['fecha_hasta'] = form.cleaned_data['fecha_hasta']
        
        if form.cleaned_data.get('usuario'):
            filtros['usuario'] = form.cleaned_data['usuario'].idusuario
        
        if form.cleaned_data.get('accion'):
            filtros['accion'] = form.cleaned_data['accion']
        
        if form.cleaned_data.get('modulo'):
            filtros['modulo'] = form.cleaned_data['modulo']
        
        # Obtener datos de auditoría (por implementar completamente)
        # Por ahora, datos de ejemplo
        data = []
        stats = {
            'total_actividades': 0,
            'usuarios_activos': 0,
            'modulos_utilizados': 0
        }
        
        # Registrar generación del reporte
        end_time = time.time()
        tiempo_generacion = end_time - start_time
        
        usuario_actual = getattr(request, 'nexo_user', None)
        try:
            ReportService.log_report_generation(
                usuario=usuario_actual,
                tipo_reporte='auditoria_sistema',
                formato=form.cleaned_data.get('formato_exportacion', 'view'),
                filtros=filtros,
                total_registros=len(data),
                tiempo_generacion=tiempo_generacion,
                ip_address=get_client_ip(request)
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar el log del reporte: {str(e)}")
        
        # Manejar exportación
        if form.cleaned_data.get('formato_exportacion') == 'pdf':
            return export_to_pdf(request, 'Auditoría del Sistema', data, stats)
        elif form.cleaned_data.get('formato_exportacion') == 'excel':
            return export_to_excel(request, 'Auditoría del Sistema', data, stats)
    
    # Paginación
    paginator = Paginator(data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    usuario_actual = getattr(request, 'nexo_user', None)
    user_iniciales = usuario_actual.nombreusuario[:2].upper() if usuario_actual and usuario_actual.nombreusuario else "IN"
    nexo_user_role = usuario_actual.rol if usuario_actual and usuario_actual.rol else 'Usuario'
    
    context = {
        'form': form,
        'page_obj': page_obj,
        'stats': stats,
        'page_title': 'Reporte de Auditoría del Sistema',
        'page_subtitle': 'Eventos clave: registros, ediciones, eliminaciones',
        'usuario_actual': usuario_actual,
        'user_iniciales': user_iniciales,
        'nexo_user_role': nexo_user_role,
        'colores': COLORES_NEXO,
    }
    
    return render(request, 'Informes/auditoria.html', context)

def export_to_pdf(request, titulo, data, stats):
    """
    Exporta reporte a PDF
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        
        # Crear buffer
        buffer = BytesIO()
        
        # Crear documento PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para el título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#4ECDC4')
        )
        
        # Encabezado del reporte
        elements.append(Paragraph("Sistema de Inventario y Ventas NEXO", title_style))
        elements.append(Paragraph("Sunshine Handmade", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        # Información del reporte
        usuario_actual = getattr(request, 'nexo_user', None)
        info_reporte = f"""
        <b>Reporte:</b> {titulo}<br/>
        <b>Usuario:</b> {usuario_actual.nombreusuario if usuario_actual else 'N/A'}<br/>
        <b>Fecha de generación:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}<br/>
        <b>Total de registros:</b> {len(data)}
        """
        elements.append(Paragraph(info_reporte, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Resumen ejecutivo si hay estadísticas
        if stats:
            elements.append(Paragraph("Resumen Ejecutivo", styles['Heading2']))
            stats_text = ""
            for key, value in stats.items():
                stats_text += f"<b>{key.replace('_', ' ').title()}:</b> {value}<br/>"
            elements.append(Paragraph(stats_text, styles['Normal']))
            elements.append(Spacer(1, 20))
        
        # Tabla de datos
        if data:
            elements.append(Paragraph("Datos del Reporte", styles['Heading2']))
            
            # Preparar datos para la tabla
            if data:
                headers = list(data[0].keys())
                table_data = [headers]
                
                for row in data[:50]:  # Limitar a 50 registros para PDF
                    table_data.append([str(row.get(col, '')) for col in headers])
                
                # Crear tabla
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4ECDC4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{titulo.replace(" ", "_")}.pdf"'
        
        return response
        
    except ImportError:
        messages.error(request, 'La exportación a PDF no está disponible. Instale reportlab.')
        return redirect(request.META.get('HTTP_REFERER', 'Informes:Informes_dashboard'))
    except Exception as e:
        logger.error(f"Error al exportar PDF: {str(e)}")
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect(request.META.get('HTTP_REFERER', 'Informes:Informes_dashboard'))

def export_to_excel(request, titulo, data, stats):
    """
    Exporta reporte a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:31]  # Excel limita nombres de hojas a 31 caracteres
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Encabezado del reporte
        ws['A1'] = "Sistema de Inventario y Ventas NEXO - Sunshine Handmade"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Reporte: {titulo}"
        ws['A2'].font = Font(bold=True)
        
        usuario_actual = getattr(request, 'nexo_user', None)
        ws['A3'] = f"Usuario: {usuario_actual.nombreusuario if usuario_actual else 'N/A'}"
        ws['A4'] = f"Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A5'] = f"Total de registros: {len(data)}"
        
        # Resumen ejecutivo
        row_start = 7
        if stats:
            ws[f'A{row_start}'] = "RESUMEN EJECUTIVO"
            ws[f'A{row_start}'].font = Font(bold=True, size=12)
            row_start += 1
            
            for key, value in stats.items():
                ws[f'A{row_start}'] = key.replace('_', ' ').title()
                ws[f'B{row_start}'] = str(value)
                row_start += 1
            
            row_start += 2
        
        # Datos del reporte
        if data:
            ws[f'A{row_start}'] = "DATOS DEL REPORTE"
            ws[f'A{row_start}'].font = Font(bold=True, size=12)
            row_start += 2
            
            # Convertir a DataFrame para facilitar la exportación
            df = pd.DataFrame(data)
            
            # Escribir headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=row_start, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Escribir datos
            for row_num, row_data in enumerate(df.values, row_start + 1):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=str(cell_value))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{titulo.replace(" ", "_")}.xlsx"'
        
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'La exportación a Excel no está disponible. Instale openpyxl y pandas.')
        return redirect(request.META.get('HTTP_REFERER', 'Informes:Informes_dashboard'))
    except Exception as e:
        logger.error(f"Error al exportar Excel: {str(e)}")
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect(request.META.get('HTTP_REFERER', 'Informes:Informes_dashboard'))

@require_GET
def dashboard_stats_api(request):
    """
    API para estadísticas del dashboard de reportes (SQL crudo)
    """
    stats = ReportService.get_dashboard_stats()
    context = {
        # ...otros datos...
        'stats': stats,
    }
    return JsonResponse({'stats': stats})
