from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from django.utils import timezone
from decimal import Decimal


def format_datetime_managua(value, include_seconds=False):
    if not value:
        return 'No especificada'
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    formato = '%d/%m/%Y %I:%M:%S %p' if include_seconds else '%d/%m/%Y %I:%M %p'
    return value.strftime(formato)

class ExportadorInformes:
    """
    Clase utilitaria mejorada para exportar informes a PDF y Excel con gráficos
    """
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """
        Configurar estilos personalizados para los PDFs
        """
        # Estilo para el título principal
        self.styles.add(ParagraphStyle(
            name='TituloNexo',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            textColor=colors.HexColor('#39bfb2'),
            alignment=1,  # Centrado
            fontName='Helvetica-Bold'
        ))
        
        # Estilo para subtítulos
        self.styles.add(ParagraphStyle(
            name='SubtituloNexo',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=15,
            textColor=colors.HexColor('#F28627'),
            fontName='Helvetica-Bold'
        ))
        
        # Estilo para información del reporte
        self.styles.add(ParagraphStyle(
            name='InfoReporte',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            fontName='Helvetica'
        ))
    
    def crear_encabezado_pdf(self, titulo, usuario, fecha_generacion, info_adicional=None):
        """
        Crear encabezado estándar para todos los PDFs
        """
        story = []
        
        # Logo y título principal
        story.append(Paragraph("NEXO – Sistema de Gestión de Inventario", self.styles['TituloNexo']))
        story.append(Paragraph(titulo, self.styles['SubtituloNexo']))
        story.append(Spacer(1, 20))
        
        # Información del reporte
        info_texto = f"""
        <b>Usuario que genera:</b> {usuario}<br/>
        <b>Fecha y hora de generación:</b> {format_datetime_managua(fecha_generacion, include_seconds=True)}<br/>
        """
        
        if info_adicional:
            for key, value in info_adicional.items():
                info_texto += f"<b>{key}:</b> {value}<br/>"
        
        story.append(Paragraph(info_texto, self.styles['InfoReporte']))
        story.append(Spacer(1, 30))
        
        return story
    
    def crear_tabla_resumen(self, datos_resumen):
        """
        Crear tabla de resumen con métricas clave
        """
        data = [['Métrica', 'Valor']]
        
        for key, value in datos_resumen.items():
            if isinstance(value, Decimal):
                value = f"C$ {value:,.2f}"
            elif isinstance(value, (int, float)):
                value = f"{value:,}"
            data.append([key.replace('_', ' ').title(), str(value)])
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        return table
    
    def generar_pdf_inventario(self, datos, usuario):
        """
        Generar PDF mejorado para el informe de inventario general
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Total de productos': datos['resumen']['total_productos'],
            'Total de existencias': datos['resumen']['total_existencias'],
            'Productos con stock bajo': datos['resumen']['productos_bajo_stock'],
            'Valor total del inventario': f"C$ {datos['resumen']['valor_total']:,.2f}"
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Inventario General", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen Ejecutivo", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de productos
        story.append(Paragraph("Detalle de Productos", self.styles['SubtituloNexo']))
        
        data = [['Código', 'Producto', 'Categoría', 'Ubicación', 'Existencia', 'Precio', 'Valor Total', 'Estado']]
        
        for producto in datos['productos']:
            valor_total = producto.existenciaproducto * (producto.precioproducto or 0)
            estado = "Stock Bajo" if producto.existenciaproducto <= (producto.existenciaminima or 5) else "Normal"
            
            data.append([
                str(producto.id_producto),
                producto.nombreproducto[:30] + "..." if len(producto.nombreproducto) > 30 else producto.nombreproducto,
                producto.idcategoriapro.nombrecategoria if producto.idcategoriapro else 'Sin categoría',
                producto.idubicacionpro.nombreubicacion,
                str(producto.existenciaproducto),
                f"C$ {producto.precioproducto or 0:.2f}",
                f"C$ {valor_total:.2f}",
                estado
            ])
        
        # Crear tabla con ancho de columnas optimizado
        table = Table(data, colWidths=[0.8*inch, 2*inch, 1.2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_pdf_ventas(self, datos, usuario, fecha_inicio, fecha_fin):
        """
        Generar PDF mejorado para el informe de ventas
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Período': f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}" if fecha_inicio and fecha_fin else "Últimos 30 días",
            'Total de ventas': datos['resumen']['total_ventas'],
            'Monto total': f"C$ {datos['resumen']['monto_total']:,.2f}",
            'Promedio por venta': f"C$ {datos['resumen']['promedio_venta']:,.2f}"
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Ventas", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen de Ventas", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de ventas
        story.append(Paragraph("Detalle de Ventas", self.styles['SubtituloNexo']))
        
        data = [['ID Venta', 'Fecha', 'Cliente', 'Vendedor', 'Total', 'Estado']]
        
        for venta in datos['ventas']:
            cliente_nombre = venta.codcliente.idpersonacliente.nombre_completo if venta.codcliente and venta.codcliente.idpersonacliente else 'Cliente no especificado'
            vendedor_nombre = venta.idusuarioventa.empleado_nombre or venta.idusuarioventa.nombreusuario if venta.idusuarioventa else 'No especificado'
            
            # Truncar nombres largos
            if len(cliente_nombre) > 25:
                cliente_nombre = cliente_nombre[:22] + "..."
            if len(vendedor_nombre) > 20:
                vendedor_nombre = vendedor_nombre[:17] + "..."
            
            data.append([
                str(venta.id_venta),
                format_datetime_managua(venta.fechaventa),
                cliente_nombre,
                vendedor_nombre,
                f"C$ {venta.total or 0:.2f}",
                venta.estado
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.8*inch, 1.2*inch, 1.8*inch, 1.5*inch, 1*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_pdf_devoluciones(self, datos, usuario, fecha_inicio, fecha_fin):
        """
        Generar PDF para el informe de devoluciones
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Período': f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}" if fecha_inicio and fecha_fin else "Todos los registros",
            'Total de devoluciones': datos['resumen']['total_devoluciones'],
            'Monto devuelto': f"C$ {datos['resumen']['monto_devuelto']:,.2f}"
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Devoluciones", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen de Devoluciones", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de devoluciones
        story.append(Paragraph("Detalle de Devoluciones", self.styles['SubtituloNexo']))
        
        data = [['Fecha', 'Producto', 'Cliente', 'Cantidad', 'Motivo', 'Monto']]
        
        for devolucion in datos['devoluciones']:
            cliente_nombre = 'No especificado'
            if devolucion.idventa and devolucion.idventa.codcliente and devolucion.idventa.codcliente.idpersonacliente:
                cliente_nombre = devolucion.idventa.codcliente.idpersonacliente.nombre_completo
            
            monto = devolucion.cantidad * (devolucion.idproducto.precioproducto or Decimal('0.00'))
            motivo = getattr(devolucion, 'motivo', 'No especificado')
            
            data.append([
                devolucion.fechadevolucion.strftime('%d/%m/%Y') if devolucion.fechadevolucion else 'No especificada',
                devolucion.idproducto.nombreproducto[:30] + "..." if len(devolucion.idproducto.nombreproducto) > 30 else devolucion.idproducto.nombreproducto,
                cliente_nombre[:25] + "..." if len(cliente_nombre) > 25 else cliente_nombre,
                str(devolucion.cantidad),
                motivo[:20] + "..." if len(motivo) > 20 else motivo,
                f"C$ {monto:.2f}"
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[1*inch, 2*inch, 1.5*inch, 0.8*inch, 1.2*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_pdf_produccion(self, datos, usuario, fecha_inicio, fecha_fin):
        """
        Generar PDF para el informe de producción
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Período': f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}" if fecha_inicio and fecha_fin else "Últimos 30 días",
            'Total de producciones': datos['resumen']['total_producciones'],
            'Cantidad total': datos['resumen']['cantidad_total'],
            'Costo total': f"C$ {datos['resumen']['costo_total']:,.2f}"
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Producción", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen de Producción", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de producciones
        story.append(Paragraph("Detalle de Producciones", self.styles['SubtituloNexo']))
        
        data = [['Fecha', 'Producto', 'Cantidad', 'Costo', 'Estado']]
        
        for produccion in datos['producciones']:
            costo = getattr(produccion, 'costo', Decimal('0.00')) or Decimal('0.00')
            
            data.append([
                produccion.fechaentrada.strftime('%d/%m/%Y') if produccion.fechaentrada else 'No especificada',
                produccion.idproducto.nombreproducto[:40] + "..." if len(produccion.idproducto.nombreproducto) > 40 else produccion.idproducto.nombreproducto,
                str(produccion.cantidadproducida),
                f"C$ {costo:.2f}",
                'Activo' if produccion.estadoregistro else 'Inactivo'
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[1.2*inch, 2.5*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_pdf_clientes(self, datos, usuario):
        """
        Generar PDF para el informe de clientes
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Total de clientes': datos['resumen']['total_clientes'],
            'Clientes activos': datos['resumen']['clientes_activos'],
            'Clientes inactivos': datos['resumen']['clientes_inactivos'],
            'Porcentaje activos': f"{datos['resumen']['porcentaje_activos']}%"
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Clientes", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen de Clientes", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de clientes
        story.append(Paragraph("Detalle de Clientes", self.styles['SubtituloNexo']))
        
        data = [['Código', 'Nombre', 'Teléfono', 'Email', 'Estado']]
        
        for cliente in datos['clientes']:
            persona = cliente.idpersonacliente
            nombre_completo = persona.nombre_completo if persona else 'No especificado'
            telefono = getattr(persona, 'telefono', 'No especificado') if persona else 'No especificado'
            email = getattr(persona, 'email', 'No especificado') if persona else 'No especificado'
            
            data.append([
                str(cliente.id_cliente),
                nombre_completo[:30] + "..." if len(nombre_completo) > 30 else nombre_completo,
                telefono,
                email[:25] + "..." if len(email) > 25 else email,
                'Activo' if cliente.estadocliente else 'Inactivo'
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.8*inch, 2.2*inch, 1.2*inch, 1.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_pdf_usuarios_empleados(self, datos, usuario):
        """
        Generar PDF para el informe de usuarios y empleados
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Total de usuarios': datos['resumen']['total_usuarios'],
            'Usuarios activos': datos['resumen']['usuarios_activos'],
            'Usuarios inactivos': datos['resumen']['usuarios_inactivos']
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Usuarios y Empleados", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen de Usuarios", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de usuarios
        story.append(Paragraph("Detalle de Usuarios", self.styles['SubtituloNexo']))
        
        data = [['Usuario', 'Nombre Empleado', 'Rol', 'Estado', 'Último Acceso']]
        
        for usuario_item in datos['usuarios']:
            ultimo_acceso = getattr(usuario_item, 'last_login', 'Nunca')
            if ultimo_acceso and ultimo_acceso != 'Nunca':
                ultimo_acceso = ultimo_acceso.strftime('%d/%m/%Y')
            
            data.append([
                usuario_item.nombreusuario,
                usuario_item.empleado_nombre or 'No especificado',
                usuario_item.rol or 'Sin rol',
                'Activo' if usuario_item.activo else 'Inactivo',
                ultimo_acceso
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[1.5*inch, 2*inch, 1.2*inch, 0.8*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_pdf_productos_categoria(self, datos, usuario):
        """
        Generar PDF para el informe de productos por categoría
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Encabezado
        info_adicional = {
            'Total de categorías': datos['resumen']['total_categorias'],
            'Total de productos': datos['resumen']['total_productos'],
            'Categoría mayor': datos['resumen']['categoria_mayor'],
            'Valor total inventario': f"C$ {datos['resumen']['valor_total_inventario']:,.2f}"
        }
        
        story.extend(self.crear_encabezado_pdf(
            "Reporte de Productos por Categoría", 
            usuario.empleado_nombre or usuario.nombreusuario,
            timezone.now(),
            info_adicional
        ))
        
        # Tabla de resumen
        story.append(Paragraph("Resumen por Categorías", self.styles['SubtituloNexo']))
        story.append(self.crear_tabla_resumen(datos['resumen']))
        story.append(Spacer(1, 20))
        
        # Tabla de categorías
        story.append(Paragraph("Detalle por Categorías", self.styles['SubtituloNexo']))
        
        data = [['Categoría', 'Cantidad Productos', 'Valor Total', 'Porcentaje']]
        
        for categoria_data in datos['categorias']:
            data.append([
                categoria_data['categoria'].nombrecategoria,
                str(categoria_data['cantidad_productos']),
                f"C$ {categoria_data['valor_total']:,.2f}",
                f"{categoria_data['porcentaje']}%"
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39bfb2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    # Métodos para generar Excel
    def generar_excel_inventario(self, datos):
        """
        Generar Excel mejorado para el informe de inventario
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario General"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = "NEXO - Reporte de Inventario General"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN EJECUTIVO"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de productos', datos['resumen']['total_productos']],
            ['Total de existencias', datos['resumen']['total_existencias']],
            ['Productos con stock bajo', datos['resumen']['productos_bajo_stock']],
            ['Valor total del inventario', f"C$ {datos['resumen']['valor_total']:,.2f}"],
            ['Porcentaje stock bajo', f"{datos['resumen']['porcentaje_stock_bajo']}%"]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['Código', 'Producto', 'Categoría', 'Ubicación', 'Existencia', 'Precio', 'Valor Total', 'Estado']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de productos
        for row, producto in enumerate(datos['productos'], start=start_row + 1):
            valor_total = producto.existenciaproducto * (producto.precioproducto or 0)
            estado = "Stock Bajo" if producto.existenciaproducto <= (producto.existenciaminima or 5) else "Normal"
            
            data_row = [
                producto.id_producto,
                producto.nombreproducto,
                producto.idcategoriapro.nombrecategoria if producto.idcategoriapro else 'Sin categoría',
                producto.idubicacionpro.nombreubicacion,
                producto.existenciaproducto,
                producto.precioproducto or 0,
                valor_total,
                estado
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [5, 6, 7]:  # Columnas numéricas
                    cell.alignment = Alignment(horizontal='right')
                    if col in [6, 7]:  # Precios
                        cell.number_format = '"C$"#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = [10, 30, 20, 15, 12, 12, 15, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_excel_ventas(self, datos):
        """
        Generar Excel para el informe de ventas
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "NEXO - Reporte de Ventas"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN DE VENTAS"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de ventas', datos['resumen']['total_ventas']],
            ['Monto total', f"C$ {datos['resumen']['monto_total']:,.2f}"],
            ['Promedio por venta', f"C$ {datos['resumen']['promedio_venta']:,.2f}"],
            ['Venta máxima', f"C$ {datos['resumen']['venta_maxima']:,.2f}"],
            ['Venta mínima', f"C$ {datos['resumen']['venta_minima']:,.2f}"]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['ID Venta', 'Fecha', 'Cliente', 'Vendedor', 'Total', 'Estado']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de ventas
        for row, venta in enumerate(datos['ventas'], start=start_row + 1):
            cliente_nombre = venta.codcliente.idpersonacliente.nombre_completo if venta.codcliente and venta.codcliente.idpersonacliente else 'Cliente no especificado'
            vendedor_nombre = venta.idusuarioventa.empleado_nombre or venta.idusuarioventa.nombreusuario if venta.idusuarioventa else 'No especificado'
            
            data_row = [
                venta.id_venta,
                format_datetime_managua(venta.fechaventa),
                cliente_nombre,
                vendedor_nombre,
                venta.total or 0,
                venta.estado
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 5:  # Columna de total
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '"C$"#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = [12, 18, 25, 20, 15, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_excel_devoluciones(self, datos):
        """
        Generar Excel para el informe de devoluciones
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Devoluciones"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "NEXO - Reporte de Devoluciones"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN DE DEVOLUCIONES"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de devoluciones', datos['resumen']['total_devoluciones']],
            ['Monto devuelto', f"C${datos['resumen']['monto_devuelto']:,.2f}"]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['Fecha', 'Producto', 'Cliente', 'Cantidad', 'Motivo', 'Monto']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de devoluciones
        for row, devolucion in enumerate(datos['devoluciones'], start=start_row + 1):
            cliente_nombre = 'No especificado'
            if devolucion.idventa and devolucion.idventa.codcliente and devolucion.idventa.codcliente.idpersonacliente:
                cliente_nombre = devolucion.idventa.codcliente.idpersonacliente.nombre_completo
            
            monto = devolucion.cantidad * (devolucion.idproducto.precioproducto or Decimal('0.00'))
            motivo = getattr(devolucion, 'motivo', 'No especificado')
            
            data_row = [
                devolucion.fechadevolucion.strftime('%d/%m/%Y') if devolucion.fechadevolucion else 'No especificada',
                devolucion.idproducto.nombreproducto,
                cliente_nombre,
                devolucion.cantidad,
                motivo,
                float(monto)
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 6:  # Columna de monto
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '"C$"#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = [15, 30, 25, 12, 20, 15]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_excel_produccion(self, datos):
        """
        Generar Excel para el informe de producción
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Producción"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = "NEXO - Reporte de Producción"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN DE PRODUCCIÓN"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de producciones', datos['resumen']['total_producciones']],
            ['Cantidad total', datos['resumen']['cantidad_total']],
            ['Costo total', f"C${datos['resumen']['costo_total']:,.2f}"],
            ['Promedio diario', datos['resumen']['promedio_diario']]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['Fecha', 'Producto', 'Cantidad', 'Costo', 'Estado']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de producción
        for row, produccion in enumerate(datos['producciones'], start=start_row + 1):
            costo = getattr(produccion, 'costo', Decimal('0.00')) or Decimal('0.00')
            
            data_row = [
                produccion.fechaentrada.strftime('%d/%m/%Y') if produccion.fechaentrada else 'No especificada',
                produccion.idproducto.nombreproducto,
                produccion.cantidadproducida,
                float(costo),
                'Activo' if produccion.estadoregistro else 'Inactivo'
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 4:  # Columna de costo
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '"C$"#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = [15, 35, 12, 15, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_excel_clientes(self, datos):
        """
        Generar Excel para el informe de clientes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Clientes"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = "NEXO - Reporte de Clientes"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN DE CLIENTES"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de clientes', datos['resumen']['total_clientes']],
            ['Clientes activos', datos['resumen']['clientes_activos']],
            ['Clientes inactivos', datos['resumen']['clientes_inactivos']],
            ['Porcentaje activos', f"{datos['resumen']['porcentaje_activos']}%"]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['Código', 'Nombre', 'Teléfono', 'Email', 'Estado']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de clientes
        for row, cliente in enumerate(datos['clientes'], start=start_row + 1):
            persona = cliente.idpersonacliente
            nombre_completo = persona.nombre_completo if persona else 'No especificado'
            telefono = getattr(persona, 'telefono', 'No especificado') if persona else 'No especificado'
            email = getattr(persona, 'email', 'No especificado') if persona else 'No especificado'
            
            data_row = [
                cliente.id_cliente,
                nombre_completo,
                telefono,
                email,
                'Activo' if cliente.estadocliente else 'Inactivo'
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = [12, 30, 18, 25, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_excel_usuarios_empleados(self, datos):
        """
        Generar Excel para el informe de usuarios y empleados
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios y Empleados"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = "NEXO - Reporte de Usuarios y Empleados"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN DE USUARIOS"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de usuarios', datos['resumen']['total_usuarios']],
            ['Usuarios activos', datos['resumen']['usuarios_activos']],
            ['Usuarios inactivos', datos['resumen']['usuarios_inactivos']]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['Usuario', 'Nombre Empleado', 'Rol', 'Estado', 'Último Acceso']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de usuarios
        for row, usuario in enumerate(datos['usuarios'], start=start_row + 1):
            ultimo_acceso = getattr(usuario, 'last_login', 'Nunca')
            if ultimo_acceso and ultimo_acceso != 'Nunca':
                ultimo_acceso = ultimo_acceso.strftime('%d/%m/%Y')
            
            data_row = [
                usuario.nombreusuario,
                usuario.empleado_nombre or 'No especificado',
                usuario.rol or 'Sin rol',
                'Activo' if usuario.activo else 'Inactivo',
                ultimo_acceso
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = [20, 25, 18, 12, 18]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generar_excel_productos_categoria(self, datos):
        """
        Generar Excel para el informe de productos por categoría
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos por Categoría"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="39BFB2", end_color="39BFB2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = "NEXO - Reporte de Productos por Categoría"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Generado el: {format_datetime_managua(timezone.now(), include_seconds=True)}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Resumen
        ws['A4'] = "RESUMEN POR CATEGORÍAS"
        ws['A4'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ['Total de categorías', datos['resumen']['total_categorias']],
            ['Total de productos', datos['resumen']['total_productos']],
            ['Categoría mayor', datos['resumen']['categoria_mayor']],
            ['Valor total inventario', f"C${datos['resumen']['valor_total_inventario']:,.2f}"]
        ]
        
        for i, (key, value) in enumerate(resumen_data, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Encabezados de la tabla principal
        headers = ['Categoría', 'Cantidad Productos', 'Valor Total', 'Porcentaje']
        start_row = len(resumen_data) + 7
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de categorías
        for row, categoria_data in enumerate(datos['categorias'], start=start_row + 1):
            data_row = [
                categoria_data['categoria'].nombrecategoria,
                categoria_data['cantidad_productos'],
                float(categoria_data['valor_total']),
                categoria_data['porcentaje']
            ]
            
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 3:  # Columna de valor total
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '"C$"#,##0.00'
                elif col == 4:  # Columna de porcentaje
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '0.00"%"'
        
        # Ajustar ancho de columnas
        column_widths = [25, 18, 18, 15]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
